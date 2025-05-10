import pandas as pd
import os
from openpyxl import load_workbook

def get_sheet_names(excel_file):
    try:
        wb = load_workbook(excel_file, read_only=True)
        return wb.sheetnames
    except Exception as e:
        print(f"Error reading sheet names: {e}")
        return []

def read_excel_sheet(excel_file, sheet_name):
    try:
        return pd.read_excel(excel_file, sheet_name=sheet_name)
    except Exception as e:
        print(f"Error reading sheet {sheet_name}: {e}")
        return None

def analyze_excel_file(excel_file):
    print(f"\nAnalyzing Excel file: {excel_file}")
    
    # Get sheet names
    sheet_names = get_sheet_names(excel_file)
    print(f"\nFound {len(sheet_names)} sheets: {', '.join(sheet_names)}")
    
    for sheet in sheet_names:
        print(f"\n--- Sheet: {sheet} ---")
        df = read_excel_sheet(excel_file, sheet)
        
        if df is not None:
            print(f"Dimensions: {df.shape[0]} rows × {df.shape[1]} columns")
            
            # Display column names
            print("\nColumns:")
            for col in df.columns:
                print(f"  - {col}")
            
            # Display first few rows
            print("\nPreview of data:")
            print(df.head(5).to_string())
        else:
            print("Could not read sheet data")

# Main
if __name__ == "__main__":
    excel_file = "attached_assets/dream eCTD machine 2.xlsx"
    if os.path.exists(excel_file):
        analyze_excel_file(excel_file)
    else:
        print(f"File not found: {excel_file}")